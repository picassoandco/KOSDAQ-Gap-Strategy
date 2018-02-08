import pandas as pd
import datetime
import numpy as np
import xlsxwriter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font,Color

def yesterdays_profit():
    file_location = "TradeLog.xlsx"
    xl = pd.ExcelFile(file_location)
    df=xl.parse("Sheet1")


    return (df.iloc[-2, -11])

def main():

    todays_date=datetime.datetime.today().date()
    print(todays_date)

    wb = Workbook()

    # 읽을 파일을 열기
    wb = load_workbook("TradeLog.xlsx")


    # 현재 actice한 sheet에 쓰기
    ws = wb.active

    # 쓸 항목을 리스트에 넣어서 ,로 구분
    ws.append([0.1, todays_date, 232323,"Asdasd2"])

    number_of_row="32"
    #퍼세티지 포맷
    ws['A'+number_of_row].number_format = "0%"
    #색깔 입히기
    ws['A' + number_of_row].font=Font(color=colors.RED)
    # 달러표시, 원화는 이상하게 계속 안되네
    ws['C'+number_of_row].number_format="0.00$"



    # 저장할 파일 선택
    wb.save("sample.xlsx")


if __name__=="__main__":
    main()
    print(yesterdays_profit())